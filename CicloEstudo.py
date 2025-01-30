import tkinter as tk
from tkinter import messagebox, simpledialog
import os
from openpyxl import Workbook, load_workbook
import winsound
import threading
import time  # Importa o módulo time para usar time.strftime()

# Lista de matérias
materias = []
index_materia = 0
tempo_limite = 9000  # 2h30min = 9000 segundos
contador = 0
pausado = False

# Função para tocar o alarme em um thread separado
def tocar_alarme():
    def tocar():
        winsound.PlaySound("alarme.wav", winsound.SND_FILENAME)
    alarme_thread = threading.Thread(target=tocar)
    alarme_thread.start()

# Função para carregar as matérias do arquivo materias.txt
def carregar_materias_txt():
    if os.path.exists("materias.txt"):
        with open("materias.txt", "r", encoding="utf-8") as file:
            materias.clear()  # Limpa a lista de matérias
            for line in file:
                materia = line.strip()  # Remove espaços extras
                if materia:  # Verifica se não está vazio
                    materias.append(materia)
    if not materias:
        # Se o arquivo estiver vazio ou não existir, usa as matérias padrão
        materias.extend(["Português", "Matemática", "História", "Física", "Geografia", "Química", "Redação", "Inglês"])
        salvar_materias_txt()  # Cria o arquivo com as matérias padrão

# Função para salvar as matérias no arquivo materias.txt
def salvar_materias_txt():
    with open("materias.txt", "w", encoding="utf-8") as file:
        for materia in materias:
            file.write(materia + "\n")

# Função para carregar a última matéria estudada
def carregar_ultima_materia():
    global index_materia
    if os.path.exists("progresso_estudos.xlsx"):
        wb = load_workbook("progresso_estudos.xlsx")
        sheet = wb.active
        ultima_linha = sheet.max_row
        
        if ultima_linha > 1:  # Se houver matérias registradas
            ultima_materia = sheet.cell(row=ultima_linha, column=2).value  # Coluna B
            if ultima_materia in materias:
                index_materia = materias.index(ultima_materia) + 1  # Avança para a próxima matéria
                if index_materia >= len(materias):  # Se for a última matéria, volta para o começo
                    index_materia = 0
            else:
                index_materia = 0  # Caso a última matéria não esteja na lista, começa pela primeira
        else:
            index_materia = 0  # Caso o arquivo esteja vazio ou sem matérias registradas, começa pela primeira matéria

# Função para criar o arquivo Excel caso não exista
def criar_arquivo():
    if not os.path.exists("progresso_estudos.xlsx"):
        wb = Workbook()
        sheet = wb.active
        # Cria os cabeçalhos
        sheet["A1"] = "Data e Hora"
        sheet["B1"] = "Matéria"
        sheet["C1"] = "Tempo Estudado"
        sheet["D1"] = "Anotações"
        wb.save("progresso_estudos.xlsx")

# Função para salvar progresso (melhorada)
def salvar_progresso():
    anotacoes = simpledialog.askstring("Anotações", "Digite suas anotações para esta matéria:")
    if anotacoes is None:  # Verifica se o usuário cancelou a caixa de diálogo
        return  # Sai da função se o usuário cancelou

    try:
        wb = load_workbook("progresso_estudos.xlsx")
        sheet = wb.active
    except FileNotFoundError:  # Lida com o erro se o arquivo não existir
        criar_arquivo()  # Cria o arquivo se não existir
        wb = load_workbook("progresso_estudos.xlsx")
        sheet = wb.active

    data_e_hora = time.strftime("%Y-%m-%d %H:%M")
    nova_linha = [data_e_hora, materias[index_materia], f"{contador // 60} min", anotacoes if anotacoes else "Nenhuma anotação"]

    proxima_linha = sheet.max_row + 1
    for col_num, value in enumerate(nova_linha, 1):
        sheet.cell(row=proxima_linha, column=col_num).value = value

    wb.save("progresso_estudos.xlsx")

# Função para carregar o tempo salvo
def carregar_tempo():
    global contador
    if os.path.exists("tempo.txt"):
        with open("tempo.txt", "r") as file:
            try:
                contador = int(file.read().strip())
            except ValueError:
                contador = 0  # Se houver erro no arquivo, zera o tempo

# Função para zerar o tempo
def zerar_tempo():
    global contador
    contador = 0
    tempo_label.config(text="Tempo: 0h 0m 0s")  # Atualiza a label do tempo
    messagebox.showinfo("Tempo Zerado", "O tempo de estudo foi zerado!")

# Função para salvar o tempo periodicamente
def salvar_tempo():
    with open("tempo.txt", "w") as file:
        file.write(str(contador))

# Função para pausar ou continuar
def pausar_continuar():
    global pausado
    pausado = not pausado
    botao_pausar.config(text="Continuar" if pausado else "Pausar")

# Função para avançar matéria (melhorada)
def avancar_materia():
    global index_materia, contador, pausado
    if not pausado:  # Salva o progresso apenas se o tempo não estiver pausado
        salvar_progresso()
    contador = 0
    pausado = False
    botao_pausar.config(text="Pausar")    
   # Avança para a próxima matéria, pulando a atual
    index_materia = (index_materia + 1) % len(materias)
    materia_label.config(text=f"Matéria Atual: {materias[index_materia]}")  # Atualiza a label da matéria
    tempo_label.config(text="Tempo: 0h 0m 0s")

# Função para remover matéria
def remover_materia():
    global materias
    try:
        selected = lista_materias.curselection()
        if selected:
            materia_remover = lista_materias.get(selected)
            materias.remove(materia_remover)
            lista_materias.delete(selected)
            salvar_materias_txt()  # Atualiza o arquivo de texto após remoção
            messagebox.showinfo("Sucesso", f"Matéria '{materia_remover}' removida.")
        else:
            messagebox.showwarning("Seleção inválida", "Selecione uma matéria para remover.")
    except Exception as e:
        messagebox.showerror("Erro", str(e))

# Função para adicionar matéria
def adicionar_materia():
    materia_nova = simpledialog.askstring("Adicionar Matéria", "Digite o nome da nova matéria:")
    if materia_nova:
        if materia_nova not in materias:
            materias.append(materia_nova)
            lista_materias.insert(tk.END, materia_nova)
            salvar_materias_txt()  # Atualiza o arquivo de texto com a nova lista de matérias
            messagebox.showinfo("Sucesso", f"Matéria '{materia_nova}' adicionada.")
        else:
            messagebox.showwarning("Aviso", "Essa matéria já está na lista.")

# Função para mover a matéria para cima (corrigida)
def mover_para_cima():
    selected = lista_materias.curselection()
    if selected:
        index = selected[0]
        if index > 0:
            # Remove o item da posição atual
            item = lista_materias.get(index)
            lista_materias.delete(index)
            materias.pop(index)

            # Insere o item na posição acima
            lista_materias.insert(index - 1, item)
            materias.insert(index - 1, item)

            # Atualiza a seleção para o item movido
            lista_materias.selection_clear(0, tk.END)
            lista_materias.select_set(index - 1)

            salvar_materias_txt()
        else:
            messagebox.showinfo("Aviso", "Já está na primeira posição.")
    else:
        messagebox.showwarning("Seleção inválida", "Selecione uma matéria para mover.")

# Função para mover a matéria para baixo (corrigida)
def mover_para_baixo():
    selected = lista_materias.curselection()
    if selected:
        index = selected[0]
        if index < len(materias) - 1:
            # Remove o item da posição atual
            item = lista_materias.get(index)
            lista_materias.delete(index)
            materias.pop(index)

            # Insere o item na posição abaixo
            lista_materias.insert(index + 1, item)
            materias.insert(index + 1, item)

            # Atualiza a seleção para o item movido
            lista_materias.selection_clear(0, tk.END)
            lista_materias.select_set(index + 1)

            salvar_materias_txt()
        else:
            messagebox.showinfo("Aviso", "Já está na última posição.")
    else:
        messagebox.showwarning("Seleção inválida", "Selecione uma matéria para mover.")

# Função para expandir/colapsar o painel de gerenciamento de matérias
def expandir_painel():
    if painel_materias.winfo_ismapped():
        painel_materias.pack_forget()  # Esconde o painel
        botao_expandir.config(text="Expandir Matérias")
    else:
        painel_materias.pack(pady=10)  # Exibe o painel
        botao_expandir.config(text="Recolher Matérias")

# Função para atualizar o tempo
def atualizar_tempo():
    global contador, pausado  # Declare 'pausado' como global
    if not pausado:  # Se o tempo não estiver pausado, atualiza o contador
        contador += 1
        tempo_label.config(text=f"Tempo: {contador // 3600}h {((contador % 3600) // 60)}m {contador % 60}s")
        
        # Verifica se atingiu o limite de 2h30min (9000 segundos)
        if contador >= 9000:
            pausado = True
            tocar_alarme()  # Toca o alarme em outro thread
            messagebox.showinfo("Limite de Tempo", "Você atingiu o tempo máximo de estudo!")
            
    salvar_tempo()  # Salva o tempo periodicamente
    root.after(1000, atualizar_tempo)  # Chama a função novamente após 1 segundo

# Criar interface gráfica
root = tk.Tk()
root.title("Ciclo Estudo")
root.geometry("450x550")  # Aumenta um pouco o tamanho da janela
root.iconbitmap("icone.ico")
root.config(bg="#f0f0f0")

# Carregar matérias e última matéria estudada
carregar_materias_txt()
criar_arquivo()
carregar_ultima_materia()  # Carregar a última matéria
carregar_tempo()  # Carregar o tempo salvo

# Mostrar a matéria atual
materia_label = tk.Label(root, text=f"Matéria Atual: {materias[index_materia]}", font=("Arial", 14), bg="#f0f0f0")
materia_label.pack(pady=10)

tempo_label = tk.Label(root, text="Tempo: 0h 0m 0s", font=("Arial", 12), bg="#f0f0f0")
tempo_label.pack()

botao_pausar = tk.Button(root, text="Pausar", command=pausar_continuar, font=("Arial", 12), bg="#4CAF50", fg="white")
botao_pausar.pack(pady=5)

botao_avancar = tk.Button(root, text="Avançar Matéria", command=avancar_materia, font=("Arial", 12), bg="#008CBA", fg="white")
botao_avancar.pack(pady=5)

botao_zerar = tk.Button(root, text="Zerar Tempo", command=zerar_tempo, font=("Arial", 12), bg="#f44336", fg="white")
botao_zerar.pack(pady=5)

# Painel de gerenciamento de matérias
painel_materias = tk.Frame(root, bg="#f0f0f0")

# Lista de matérias
lista_materias = tk.Listbox(painel_materias, height=5, selectmode=tk.SINGLE, font=("Arial", 12), bg="#ffffff")
for materia in materias:
    lista_materias.insert(tk.END, materia)
lista_materias.pack(pady=5)

# Botões de gerenciamento de matérias
botao_adicionar = tk.Button(painel_materias, text="Adicionar Matéria", command=adicionar_materia, font=("Arial", 12), bg="#8e44ad", fg="white")
botao_adicionar.pack(pady=5)

botao_remover = tk.Button(painel_materias, text="Remover Matéria", command=remover_materia, font=("Arial", 12), bg="#c0392b", fg="white")
botao_remover.pack(pady=5)

botao_mover_cima = tk.Button(painel_materias, text="Mover para Cima", command=mover_para_cima, font=("Arial", 12), bg="#2ecc71", fg="white")
botao_mover_cima.pack(pady=5)

botao_mover_baixo = tk.Button(painel_materias, text="Mover para Baixo", command=mover_para_baixo, font=("Arial", 12), bg="#2ecc71", fg="white")
botao_mover_baixo.pack(pady=5)

# Botão de expandir/colapsar painel
botao_expandir = tk.Button(root, text="Expandir Matérias", command=expandir_painel, font=("Arial", 12), bg="#3498db", fg="white")
botao_expandir.pack(pady=10)

# Atualiza o tempo de forma contínua
atualizar_tempo()

root.mainloop()
